import openpyxl

def filler(count):
    sheet.cell('C'+str(count+1)).value = 0
    sheet.cell('D'+str(count+1)).value = 0
    sheet.cell('E'+str(count+1)).value = 0
    sheet.cell('F'+str(count+1)).value = 0
    sheet.cell('G'+str(count+1)).value = 0
    sheet.cell('H'+str(count+1)).value = 0
    sheet.cell('I'+str(count+1)).value = 0
    sheet.cell('J'+str(count+1)).value = 0
    sheet.cell('K'+str(count+1)).value = 0
    sheet.cell('L'+str(count+1)).value = 0
    sheet.cell('M'+str(count+1)).value = 0
    sheet.cell('N'+str(count+1)).value = 0
    sheet.cell('O'+str(count+1)).value = 0
    sheet.cell('P'+str(count+1)).value = 0
    sheet.cell('Q'+str(count+1)).value = 0
    sheet.cell('R'+str(count+1)).value = 0
    sheet.cell('S'+str(count+1)).value = 0
    sheet.cell('T'+str(count+1)).value = 0
    sheet.cell('U'+str(count+1)).value = 0
    sheet.cell('V'+str(count+1)).value = 0
    sheet.cell('W'+str(count+1)).value = 0
    sheet.cell('X'+str(count+1)).value = 0

wb = openpyxl.Workbook()#load_workbook('events_dataset.xlsx')
sheet = wb.get_active_sheet()
sheet.title = "Sheet1"
file = open("view-source_www.espncricinfo.com_pakistan-v-england-2015-16_engine_match_902643.html_innings=1_view=commentary .txt", 'r')
sheet.cell('A1').value = "Serial Number"
sheet.cell('B1').value = "Discourse"
sheet.cell('C1').value = "1"
sheet.cell('D1').value = "2"
sheet.cell('E1').value = "3"
sheet.cell('F1').value = "4"
sheet.cell('G1').value = "5"
sheet.cell('H1').value = "6"
sheet.cell('I1').value = "7"
sheet.cell('J1').value = "8"
sheet.cell('K1').value = "9"
sheet.cell('L1').value = "10"
sheet.cell('M1').value = "11"
sheet.cell('N1').value = "12"
sheet.cell('O1').value = "13"
sheet.cell('P1').value = "14"
sheet.cell('Q1').value = "15"
sheet.cell('R1').value = "16"
sheet.cell('S1').value = "17"
sheet.cell('T1').value = "18"
sheet.cell('U1').value = "19"
sheet.cell('V1').value = "20"
sheet.cell('W1').value = "21"
sheet.cell('X1').value = "22"
count = 1

for line in file.readlines():
    sheet.cell('A'+str(count+1)).value = count
    sheet.cell('B'+str(count+1)).value = line  
    filler(count)
    if "FOUR" in line or "SIX" in line:
        sheet.cell('L'+str(count+1)).value = sheet.cell('L'+str(count+1)).value + 1
    if "1 run" in line or "2 runs" in line:
        sheet.cell('M'+str(count+1)).value = sheet.cell('M'+str(count+1)).value + 1


    if "run out" in line: # Run out Missed
        sheet.cell('R'+str(count+1)).value = sheet.cell('R'+str(count+1)).value + 1
    # Dropped Catch Done Manually since only 1 and confusing cases
    # No Missing Stumpings
    
    if "defended" in line:
        sheet.cell('C'+str(count+1)).value = sheet.cell('C'+str(count+1)).value + 1
    # No Left Alone (D)
    if "beaten" in line:
        sheet.cell('E'+str(count+1)).value = sheet.cell('E'+str(count+1)).value + 1
    if "edge" in line:
        sheet.cell('F'+str(count+1)).value = sheet.cell('F'+str(count+1)).value + 1
    if "1 wide" in line:
        sheet.cell('O'+str(count+1)).value = sheet.cell('O'+str(count+1)).value + 1
    # No Boundaries By Extras (N)
    
    if "yorker" in line:
        sheet.cell('T'+str(count+1)).value = sheet.cell('T'+str(count+1)).value + 1
    count = count + 1
wb.save("event_dataset_fill.xlsx")
